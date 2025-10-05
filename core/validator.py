"""
Script que permite validar mediante reglas si el archivo es xlsx
"""

from zipfile import ZipFile,BadZipFile
import subprocess
import sys
import os
from openpyxl import load_workbook

def reparar_archivo(archivo_entrada, archivo_salida):
    """
    Funcion que repara los xlsx utilizando zip -FF
    """
    print(f"Reparando {archivo_entrada} en {archivo_salida} ")
    error = False

    try:
        result = subprocess.run(
            ['zip', '-FF', f'{archivo_entrada}', '--out', f'{archivo_salida}'],
            input='y\n',
            text=True,
            capture_output=True
        )
    except Exception as e:
        error = True
        raise(f"error al reparar el archivo {archivo_entrada} : {e}")
    
    return error

def validar_archivo(archivo, carpeta_estructura, archivo_estructura, validacion="open"):

    """
    Funcion que valida si tiene la estructura de un archivo xlsx valido para la lectura
    """
    valido = False
    carpeta_presente = False
    archivo_presente = False

    if validacion == "open":

        try:
            wb = load_workbook(archivo)
            print("File is valid and can be opened.")
            return True
        
        except BadZipFile:
            print("File is not a valid ZIP archive, hence not a valid .xlsx file.")
            return False
        
        except Exception as e:
            print(f"An error occurred: {e}")
            return False
    else:

        with ZipFile(f"{archivo}", 'r') as zip_ref:

            for file_name in zip_ref.namelist():
                print(f"validando el archivo {file_name}")

                nombre = file_name.split("/")

                if nombre[0] == carpeta_estructura:
                    carpeta_presente = True
                
                if file_name == archivo_estructura:
                    archivo_presente = True

                if archivo_presente and carpeta_presente:

                    valido = True
                    break
    return valido

def validate_xlsx_file(filepath):
    try:
        wb = load_workbook(filepath)
        print("File is valid and can be opened.")
        return True
    except BadZipFile:
        print("File is not a valid ZIP archive, hence not a valid .xlsx file.")
        return False
    except Exception as e:
        print(f"An error occurred: {e}")
        return False


def main(archivo_entrada, archivo_salida, carpeta_estructura="xl", archivo_estructura="[Content_Types].xml"):
    """
    Funcion principal que repara el archivo y despues determina si es un archivo es valido
    """

    error = reparar_archivo(
        archivo_entrada=archivo_entrada,
        archivo_salida=archivo_salida
    )

    if error == False:

        valido = validar_archivo(
            archivo=archivo_salida,
            carpeta_estructura=carpeta_estructura,
            archivo_estructura=archivo_estructura
        )

        if valido == False:
            print(f"[Fail] el archivo {archivo_entrada} no es un xlsx valido")
            os.remove(f"{archivo_salida}")

if __name__ == "__main__":
    op = sys.argv[1:]
    main(op[0],op[1])
